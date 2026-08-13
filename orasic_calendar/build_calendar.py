import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Calendario 30 días"

FONT = "Arial"
DARK_BG = PatternFill("solid", fgColor="0B0E14")
HEADER_FILL = PatternFill("solid", fgColor="1E293B")
INPUT_FILL = PatternFill("solid", fgColor="FFFDE7")
WHITE = Font(name=FONT, size=11, color="FFFFFF", bold=False)
HEADER_FONT = Font(name=FONT, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=16, bold=True, color="A78BFA")
BLUE_INPUT = Font(name=FONT, size=11, bold=True, color="0000FF")

thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title
ws.merge_cells("A1:G1")
ws["A1"] = "ORASIC Lab — Calendario de contenido diario (TikTok / Instagram)"
ws["A1"].font = TITLE_FONT

ws["A2"] = "Fecha de inicio (Día 1):"
ws["A2"].font = Font(name=FONT, size=11, bold=True)
ws["B2"] = datetime.date(2026, 7, 29)
ws["B2"].font = BLUE_INPUT
ws["B2"].fill = INPUT_FILL
ws["B2"].number_format = "DD/MM/YYYY"

ws["D2"] = "Publicados:"
ws["D2"].font = Font(name=FONT, size=11, bold=True)
ws["E2"] = "=COUNTIF(G5:G34,\"Publicado\")&\" / 30\""

headers = ["Día", "Fecha", "Pilar", "Rubro / Tema", "Idea / Gancho", "Formato", "Estado"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

rows = [
 # Pilar, Rubro/Tema, Idea/Gancho, Formato
 ("Mito vs Realidad", "Costo y tiempo", "Ya publicado: “cuesta miles y tarda meses” vs. IA en días", "Narrado"),
 ("Antes/Después", "Reservas + cobro (pyme genérico)", "Ya publicado: WhatsApp desordenado vs. app que agenda y cobra", "Narrado"),
 ("Números", "+100 apps, 12 rubros, 2 países", "Ya publicado: prueba social en cifras", "Narrado"),
 ("Tips", "Señales de alerta", "3 señales de que tu negocio necesita dejar el Excel", "Silencioso"),
 ("Antes/Después", "Ventas B2B", "De hojas de cálculo con leads olvidados a un seguimiento que no se cae", "Silencioso"),
 ("Rubro del día", "Belleza", "Fichas de clientas, reservas e historial en una sola app", "Silencioso"),
 ("Detrás de cámara", "Proceso interno", "Así se ve el día 1 de un proyecto en ORASIC Lab", "Narrado"),
 ("Mito vs Realidad", "Necesito programadores propios", "Mito: hay que armar equipo técnico. Realidad: nosotros lo hacemos con IA", "Silencioso"),
 ("Antes/Después", "Retail y moda", "De un cuaderno de ventas a un dashboard que se actualiza solo", "Silencioso"),
 ("Objeciones", "Negocio pequeño", "¿Y si mi negocio es muy chico para una app?", "Narrado"),
 ("Rubro del día", "Salud", "Recordatorios de citas e historial de pacientes, sin líos de agenda", "Silencioso"),
 ("Números", "Tiempo de entrega", "De meses a días: así de rápido sale una primera versión", "Silencioso"),
 ("Tips", "Automatización con criterio", "Cómo automatizar sin perder el trato humano con tus clientes", "Narrado"),
 ("Antes/Después", "Educación", "De matrículas por WhatsApp a un seguimiento de alumnos ordenado", "Silencioso"),
 ("Mito vs Realidad", "Solo para empresas grandes", "Mito: esto es para corporativos. Realidad: pymes son la mayoría de nuestros casos", "Silencioso"),
 ("Rubro del día", "Deporte", "Reservas de canchas y clases sin depender del chat", "Silencioso"),
 ("Detrás de cámara", "Cómo priorizamos", "Cómo decidimos qué construir primero en un proyecto", "Narrado"),
 ("Antes/Después", "Marketing y contenido", "De un calendario de posts en notas sueltas a un flujo con seguimiento de leads", "Silencioso"),
 ("Objeciones", "No sé de tecnología", "¿Y si no entiendo nada de tecnología? Así te acompañamos igual", "Narrado"),
 ("Números", "Ahorro vs. agencia tradicional", "Por qué construir con IA baja el costo sin bajar la calidad", "Silencioso"),
 ("Rubro del día", "Coaching y liderazgo", "Agenda de sesiones y seguimiento de clientes en un solo lugar", "Silencioso"),
 ("Tips", "Respuesta rápida", "La señal #1 de que estás perdiendo ventas por no responder a tiempo", "Silencioso"),
 ("Antes/Después", "Productividad interna", "De tareas en chats sueltos a un flujo de trabajo que el equipo sí usa", "Silencioso"),
 ("Mito vs Realidad", "Pierdo el trato personal", "Mito: automatizar es perder cercanía. Realidad: libera tiempo para atender mejor", "Narrado"),
 ("Rubro del día", "Institucional y corporativo", "Dashboards de gestión para equipos que necesitan ver todo en un lugar", "Silencioso"),
 ("Detrás de cámara", "Qué pasa después del primer mensaje", "Qué pasa apenas nos escribes por WhatsApp", "Silencioso"),
 ("Antes/Después", "Minería y conflictos sociales", "De reportes dispersos a seguimiento centralizado de incidencias", "Silencioso"),
 ("Objeciones", "Tiempo real de entrega", "¿Cuánto se demora en verdad? Sin promesas exageradas", "Narrado"),
 ("Números", "Cierre de mes", "Resumen del mes: cifras reales, sin inventar nada", "Silencioso"),
 ("Tips", "Aprendizajes", "Lo que aprendimos construyendo más de 100 aplicaciones", "Narrado"),
]

for idx, (pilar, rubro, idea, formato) in enumerate(rows, start=1):
    r = 4 + idx
    ws.cell(row=r, column=1, value=idx).border = border
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    fc = ws.cell(row=r, column=2, value=f"=$B$2+{idx-1}")
    fc.number_format = "DD/MM/YYYY"
    fc.border = border
    ws.cell(row=r, column=3, value=pilar).border = border
    ws.cell(row=r, column=4, value=rubro).border = border
    idea_cell = ws.cell(row=r, column=5, value=idea)
    idea_cell.border = border
    idea_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(row=r, column=6, value=formato).border = border
    estado = "Publicado" if idx <= 3 else "Pendiente"
    ec = ws.cell(row=r, column=7, value=estado)
    ec.border = border
    if estado == "Publicado":
        ec.font = Font(name=FONT, color="16A34A", bold=True)
    else:
        ec.font = Font(name=FONT, color="94A3B8")

widths = {"A":8,"B":13,"C":18,"D":26,"E":50,"F":13,"G":14}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

for row in ws.iter_rows(min_row=5, max_row=34, min_col=1, max_col=7):
    for cell in row:
        if cell.font is None or cell.font.name != FONT:
            cell.font = Font(name=FONT, size=10)

ws.freeze_panes = "A5"

out = "/sessions/admiring-beautiful-pasteur/mnt/outputs/orasic_calendar/Calendario-ORASIC-Lab.xlsx"
wb.save(out)
print("saved", out)
